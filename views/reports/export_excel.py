import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


class ExcelExporter:

    # ==================================================
    # Workbook Helpers
    # ==================================================

    @staticmethod
    def create_workbook(title):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title
        return workbook, sheet

    # ==================================================
    # Styles
    # ==================================================

    @staticmethod
    def title_font():
        return Font(
            size=18,
            bold=True
        )

    @staticmethod
    def heading_font():
        return Font(
            bold=True,
            color="FFFFFF"
        )

    @staticmethod
    def heading_fill():
        return PatternFill(
            fill_type="solid",
            start_color="2563EB",
            end_color="2563EB"
        )

    @staticmethod
    def thin_border():
        side = Side(style="thin")
        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )

    @staticmethod
    def center():
        return Alignment(
            horizontal="center",
            vertical="center"
        )

    # ==================================================
    # Export Folder
    # ==================================================

    @staticmethod
    def get_export_folder():
        folder = os.path.join(
            "exports",
            "excel"
        )
        os.makedirs(
            folder,
            exist_ok=True
        )
        return folder

    # ==================================================
    # Auto Fit Columns
    # ==================================================

    @staticmethod
    def auto_fit_columns(sheet):
        for column in range(
            1,
            sheet.max_column + 1
        ):
            max_length = 0
            column_letter = get_column_letter(
                column
            )
            for row in range(
                1,
                sheet.max_row + 1
            ):
                value = sheet.cell(
                    row=row,
                    column=column
                ).value
                if value is not None:
                    max_length = max(
                        max_length,
                        len(str(value))
                    )
            sheet.column_dimensions[
                column_letter
            ].width = max_length + 5

    # ==================================================
    # Generic Report Generator
    # ==================================================

    @classmethod
    def generate_report(
        cls,
        report_name,
        last_column,
        headers,
        records,
        extractor,
        filename
    ):
        workbook, sheet = cls.create_workbook(
            report_name
        )

        #################################################
        # Title
        #################################################

        sheet.merge_cells(
            f"A1:{last_column}1"
        )
        cell = sheet["A1"]
        cell.value = "STUDENT MANAGEMENT SYSTEM"
        cell.font = cls.title_font()
        cell.alignment = cls.center()

        #################################################

        sheet.merge_cells(
            f"A2:{last_column}2"
        )
        cell = sheet["A2"]
        cell.value = f"{report_name} Report"
        cell.font = Font(
            size=14,
            bold=True
        )
        cell.alignment = cls.center()

        #################################################

        sheet["A4"] = "Export Date"
        sheet["B4"] = datetime.now().strftime(
            "%d-%m-%Y %H:%M"
        )

        #################################################
        # Headers
        #################################################

        row = 6
        for col, heading in enumerate(
            headers,
            start=1
        ):
            cell = sheet.cell(
                row=row,
                column=col
            )
            cell.value = heading
            cell.font = cls.heading_font()
            cell.fill = cls.heading_fill()
            cell.border = cls.thin_border()
            cell.alignment = cls.center()

        #################################################
        # Data
        #################################################

        row = 7
        for record in records:
            values = extractor(record)
            for col, value in enumerate(
                values,
                start=1
            ):
                cell = sheet.cell(
                    row=row,
                    column=col
                )
                cell.value = value
                cell.border = cls.thin_border()
            row += 1

        #################################################
        # Auto Width
        #################################################

        cls.auto_fit_columns(sheet)

        #################################################
        # Save
        #################################################

        filepath = os.path.join(
            cls.get_export_folder(),
            filename
        )
        workbook.save(filepath)
        return filepath

    # ==================================================
    # Export Students
    # ==================================================

    @classmethod
    def export_students(cls):
        from controllers.student_controller import StudentController

        headers = [
            "Student ID",
            "Admission No",
            "First Name",
            "Last Name",
            "Gender",
            "Course",
            "Department",
            "Semester",
            "Status"
        ]
        students = StudentController.get_all_students()

        def extractor(student):
            return [
                student.get("student_id", ""),
                student.get("admission_no", ""),
                student.get("first_name", ""),
                student.get("last_name", ""),
                student.get("gender", ""),
                student.get("course", ""),
                student.get("department", ""),
                student.get("semester", ""),
                student.get("status", "")
            ]

        return cls.generate_report(
            report_name="Students",
            last_column="I",
            headers=headers,
            records=students,
            extractor=extractor,
            filename="Students.xlsx"
        )

    # ==================================================
    # Export Faculty
    # ==================================================

    @classmethod
    def export_faculty(cls):
        from controllers.faculty_controller import FacultyController

        headers = [
            "Faculty ID",
            "First Name",
            "Last Name",
            "Gender",
            "Department",
            "Qualification",
            "Email",
            "Status"
        ]
        faculty = FacultyController.get_all_faculty()

        def extractor(member):
            return [
                member.get("faculty_id", ""),
                member.get("first_name", ""),
                member.get("last_name", ""),
                member.get("gender", ""),
                member.get("department", ""),
                member.get("qualification", ""),
                member.get("email", ""),
                member.get("status", "")
            ]

        return cls.generate_report(
            report_name="Faculty",
            last_column="H",
            headers=headers,
            records=faculty,
            extractor=extractor,
            filename="Faculty.xlsx"
        )

    # ==================================================
    # Export Courses
    # ==================================================

    @classmethod
    def export_courses(cls):
        from controllers.course_controller import CourseController

        headers = [
            "Course ID",
            "Course Name",
            "Department",
            "Semester",
            "Credits",
            "Duration",
            "Status"
        ]
        courses = CourseController.get_all_courses()

        def extractor(course):
            return [
                course.get("course_id", ""),
                course.get("course_name", ""),
                course.get("department", ""),
                course.get("semester", ""),
                course.get("credits", ""),
                course.get("duration", ""),
                course.get("status", "")
            ]

        return cls.generate_report(
            report_name="Courses",
            last_column="G",
            headers=headers,
            records=courses,
            extractor=extractor,
            filename="Courses.xlsx"
        )

    # ==================================================
    # Export Attendance
    # ==================================================

    @classmethod
    def export_attendance(cls):
        from controllers.attendance_controller import AttendanceController

        headers = [
            "Attendance ID",
            "Student ID",
            "Student Name",
            "Course",
            "Department",
            "Date",
            "Status"
        ]
        attendance = AttendanceController.get_all_attendance()

        def extractor(record):
            return [
                record.get("attendance_id", ""),
                record.get("student_id", ""),
                record.get("student_name", ""),
                record.get("course", ""),
                record.get("department", ""),
                record.get("date", ""),
                record.get("status", "")
            ]

        return cls.generate_report(
            report_name="Attendance",
            last_column="G",
            headers=headers,
            records=attendance,
            extractor=extractor,
            filename="Attendance.xlsx"
        )

    # ==================================================
    # Export Fees
    # ==================================================

    @classmethod
    def export_fees(cls):
        from controllers.fee_controller import FeeController

        headers = [
            "Fee ID",
            "Student ID",
            "Student Name",
            "Course",
            "Department",
            "Total Fee",
            "Amount Paid",
            "Balance",
            "Due Date",
            "Status"
        ]
        fees = FeeController.get_all_fees()

        def extractor(fee):
            return [
                fee.get("fee_id", ""),
                fee.get("student_id", ""),
                fee.get("student_name", ""),
                fee.get("course", ""),
                fee.get("department", ""),
                fee.get("total_fee", ""),
                fee.get("amount_paid", ""),
                fee.get("balance", ""),
                fee.get("due_date", ""),
                fee.get("status", "")
            ]

        return cls.generate_report(
            report_name="Fees",
            last_column="J",
            headers=headers,
            records=fees,
            extractor=extractor,
            filename="Fees.xlsx"
        )